"""
Python Expert
Finding_New_Addresses.py
Utilities for exporting LINZ streets (Option 10) without circular imports.
"""

from __future__ import annotations

import os
import re
import time

import pandas as pd
from tqdm import tqdm as _tqdm

__all__ = ["finding_new_addresses", "export_linz_streets", "export_linz_suburbs", "clean_exported_linz_streets"]


# --- Option 3: Clean an already-exported streets file ------------------------
def clean_exported_linz_streets(
    src_folder: str = "Street database",
    in_xlsx: str = "Exported_LINZ_Streets.xlsx",
    out_xlsx: str = "Exported_LINZ_Streets_Clean.xlsx",
) -> None:
    """
    Reads 'Streets' sheet with 'Street, Suburb', cleans, dedupes, groups A–Z,
    writes 3 sheets, and shows a progress bar for every stage.
    """
    import os, re, time
    import pandas as pd
    from contextlib import contextmanager

    @contextmanager
    def _stage_bar(desc: str, total: int = 1, unit: str = "step"):
        bar = _tqdm(total=total, desc=desc, unit=unit, dynamic_ncols=True, leave=False)
        try:
            yield bar
        finally:
            if bar.n < bar.total:
                bar.update(bar.total - bar.n)
            bar.close()

    # Leading unit/house tokens (G09/G25, 12-14A/3, A/1, SH16, etc.)
    _LEADING_PREFIX_RE = re.compile(
        r"""
        ^\s*
        (?:
            (?:Unit\s*[A-Za-z0-9]+/)?           # optional "Unit 3/"
            (?=[A-Za-z0-9/-]*\d)                # must contain a digit before the first space
            [A-Za-z0-9]+(?:[/-][A-Za-z0-9]+)*   # segments like G09/G25, 12-14A/3, B11/710
        )
        \s+                                      # spaces before street name
        """,
        re.IGNORECASE | re.VERBOSE,
    )

    def _clean_one(value: str) -> str:
        s = (value or "").strip()
        if not s or "," not in s:
            return s
        street_part, suburb_part = [p.strip() for p in s.split(",", 1)]
        street_part = _LEADING_PREFIX_RE.sub("", street_part).strip()
        street_part = " ".join(street_part.split()).title()
        suburb_part = " ".join(suburb_part.split()).title()
        return f"{street_part}, {suburb_part}"

    # 1) Locate file
    with _stage_bar("📁 Locate Excel") as bar:
        in_path = os.path.join(src_folder, in_xlsx) if os.path.isdir(src_folder) else in_xlsx
        if not os.path.exists(in_path):
            print(f"❌ Excel not found at '{in_path}'.")
            return
        bar.update(1)

    # 2) Load
    with _stage_bar("📥 Load Excel") as bar:
        try:
            df = pd.read_excel(in_path, sheet_name="Streets", dtype=str)
        except ValueError:
            df = pd.read_excel(in_path, sheet_name=0, dtype=str)
        df = df.fillna("")
        if "Street, Suburb" not in df.columns:
            print("❌ Column 'Street, Suburb' not found.")
            return
        originals = df["Street, Suburb"].astype(str).tolist()
        total_rows = len(originals)
        bar.update(1)

    # 3) Clean each row
    cleaned = []
    for v in _tqdm(originals, desc="🧹 Clean Street Names", unit="row", dynamic_ncols=True, leave=False):
        cleaned.append(_clean_one(v))

    # 4) De-duplicate + sort A→Z
    with _stage_bar("🧮 De-duplicate + Sort A→Z") as bar:
        cleaned_sorted = sorted(set(s for s in cleaned if s))
        bar.update(1)

    # 4a) Audit: originals with digits/symbols
    flagged_rows = []
    for raw, cleaned_item in _tqdm(
        list(zip(originals, cleaned)),
        total=total_rows,
        desc="🔍 Scan Originals For Numbers/Symbols",
        unit="row",
        dynamic_ncols=True,
        leave=False,
    ):
        if not raw or "," not in raw:
            continue
        street_orig = raw.split(",", 1)[0].strip()
        has_digit = any(ch.isdigit() for ch in street_orig)
        bad_symbols = sorted({ch for ch in street_orig if not (ch.isalpha() or ch in " -'" or ch.isdigit())})
        if has_digit or bad_symbols:
            issue_labels = []
            if has_digit:
                issue_labels.append("digit")
            if bad_symbols:
                issue_labels.append("symbol:" + "".join(bad_symbols))
            flagged_rows.append({
                "Original": raw.strip(),
                "Cleaned": (cleaned_item or "").strip(),
                "Street Only (Original)": street_orig,
                "Issues": ", ".join(issue_labels),
            })
    flagged_df = (
        pd.DataFrame(flagged_rows).drop_duplicates().sort_values("Original")
        if flagged_rows else
        pd.DataFrame(columns=["Original","Cleaned","Street Only (Original)","Issues"])
    )

    # 4b) Build A–Z columns
    letters = [chr(c) for c in range(ord("A"), ord("Z") + 1)]
    buckets = {L: [] for L in letters}
    buckets["Other"] = []
    for item in _tqdm(cleaned_sorted, desc="📚 Build A–Z Columns", unit="item", dynamic_ncols=True, leave=False):
        street = item.split(",", 1)[0].strip()
        ch = street[:1].upper() if street else ""
        key = ch if ch in buckets else "Other"
        buckets[key].append(item)
    max_len = max((len(v) for v in buckets.values()), default=0)
    for k in buckets:
        if len(buckets[k]) < max_len:
            buckets[k].extend([""] * (max_len - len(buckets[k])))
    include_other = any(v.strip() for v in buckets["Other"])
    ordered_cols = letters + (["Other"] if include_other else [])
    by_letter_df = pd.DataFrame({k: buckets[k] for k in ordered_cols})

    # 5) Write output (3 sheets)
    out_dir = src_folder if os.path.isdir(src_folder) else "."
    out_path = os.path.join(out_dir, out_xlsx)
    with _stage_bar("💾 Write Excel (3 sheets)", total=3, unit="sheet") as bar:
        try:
            with pd.ExcelWriter(out_path, engine="openpyxl") as xls:
                pd.DataFrame({"Street, Suburb": cleaned_sorted}).to_excel(
                    xls, sheet_name="Streets", index=False
                ); bar.update(1)
                by_letter_df.to_excel(xls, sheet_name="ByLetter", index=False); bar.update(1)
                flagged_df.to_excel(xls, sheet_name="WithNumbersOrSymbols", index=False); bar.update(1)
            print(
                f"✅ Cleaned {len(cleaned_sorted)} rows → {out_path} "
                f"(sheets: 'Streets', 'ByLetter', 'WithNumbersOrSymbols')"
            )
        except PermissionError:
            base, ext = os.path.splitext(out_xlsx)
            ts = time.strftime("%Y%m%d_%H%M%S")
            alt_path = os.path.join(out_dir, f"{base}_{ts}{ext}")
            with pd.ExcelWriter(alt_path, engine="openpyxl") as xls:
                pd.DataFrame({"Street, Suburb": cleaned_sorted}).to_excel(
                    xls, sheet_name="Streets", index=False
                ); bar.update(1)
                by_letter_df.to_excel(xls, sheet_name="ByLetter", index=False); bar.update(1)
                flagged_df.to_excel(xls, sheet_name="WithNumbersOrSymbols", index=False); bar.update(1)
            print(f"⚠️  File in use; saved as → {alt_path}")


# --- Option 4: Create Excel template for new addresses -----------------------
def create_excel_for_new_addresses(
    out_xlsx: str = "New_Addresses_Template.xlsx",
    sheet_name: str = "New Addresses",
    input_col: str = "D",
    dropdown_col: str = "E",
    first_row: int | None = None,      # default: 18
    num_rows: int | None = None,       # default: E18..E518 (501 rows)
    first_dropdown_row: int | None = None,
    num_dropdowns: int | None = None,
    block_size: int = 500,              # Z-block height per row
    source_col: str = "X",              # master list lives in X2:X50000
    source_start_row: int = 2,
    source_end_row: int = 50000,
    hide_helpers: bool = False,
) -> None:
    """
    Google Sheets–friendly setup:
      • Type in D18..D518.
      • EACH row gets its own helper Z-block of size=block_size with:
          SORT(UNIQUE(FILTER(X2:X50000, prefix matches D{row}))).
      • No data-validations are applied; you’ll paste them in Sheets UI.
    """
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Border, Side
    from contextlib import contextmanager
    import os, time as _time

    @contextmanager
    def _stage_bar(desc: str, total: int = 1, unit: str = "step"):
        bar = _tqdm(total=total, desc=desc, unit=unit, dynamic_ncols=True, leave=False)
        try:
            yield bar
        finally:
            if bar.n < bar.total:
                bar.update(bar.total - bar.n)
            bar.close()

    # Resolve target rows: default D/E18..D/E518 (501 rows)
    if first_row is None and first_dropdown_row is not None:
        first_row = first_dropdown_row
    if num_rows is None and num_dropdowns is not None:
        num_rows = num_dropdowns
    if first_row is None:
        first_row = 18
    if num_rows is None:
        num_rows = 518 - 18 + 1
    last_row = first_row + num_rows - 1

    list_col = "Z"  # helper column for per-row Z-blocks

    # 1) init workbook & layout
    with _stage_bar("🧰 Initialize workbook & layout"):
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        ws.column_dimensions[input_col].width = 28
        ws.column_dimensions[dropdown_col].width = 28
        ws.column_dimensions[source_col].width = 36
        ws.column_dimensions[list_col].width = 52
        ws[f"{input_col}1"] = f"Type HERE: {input_col}{first_row}:{input_col}{last_row}"
        ws[f"{dropdown_col}1"] = f"Dropdown HERE (manually validated in Sheets): {dropdown_col}{first_row}:{dropdown_col}{last_row}"
        ws[f"{source_col}1"] = f"Master list {source_col}{source_start_row}:{source_col}{source_end_row}"
        if hide_helpers:
            ws.column_dimensions[source_col].hidden = True
            ws.column_dimensions[list_col].hidden = True

        # styles
        block_fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
        block_font = Font(bold=True)
        edge = Side(style="thin", color="FF888888")
        dropdown_border = Border(left=edge, right=edge, top=edge, bottom=edge)

        # per-cell helper
        def _z_item_formula(row_idx: int, k: int) -> str:
            sh   = f"'{sheet_name}'"
            src  = f"{sh}!${source_col}${source_start_row}:${source_col}${source_end_row}"
            inp  = f"{sh}!${input_col}${row_idx}"
            list_expr = (
                f"SORT(UNIQUE(FILTER({src}, (LEN({src})>0) * "
                f"(LEFT(LOWER({src}), LEN(LOWER(TRIM({inp})))) = LOWER(TRIM({inp}))))))"
            )
            return (
                "=IF("
                f"LEN(TRIM({inp}))=0,"
                '""'
                f",IFERROR(INDEX({list_expr}, {k}),"
                '""'
                "))"
            )

        # Build Z-blocks per row
        for i, row_idx in enumerate(
            _tqdm(range(first_row, last_row + 1), desc="🧩 Build helpers (Z-blocks)", unit="row", dynamic_ncols=True, leave=False)
        ):
            z_top = i * block_size + 1
            z_end = z_top + block_size - 1

            z_top_cell = ws[f"{list_col}{z_top}"]
            z_top_cell.value = _z_item_formula(row_idx, 1)
            z_top_cell.fill = block_fill
            z_top_cell.font = block_font

            for k, zr in enumerate(range(z_top + 1, z_end + 1), start=2):
                ws[f"{list_col}{zr}"] = _z_item_formula(row_idx, k)

            # Optional: add a faint border to target dropdown cell so it stands out visually
            ws[f"{dropdown_col}{row_idx}"].border = dropdown_border

    # 3) save
    with _stage_bar("💾 Save Excel"):
        try:
            wb.save(out_xlsx)
            print(f"✅ Created Excel → {out_xlsx}")
            print(f"   Type only in {input_col}{first_row}:{input_col}{last_row}.")
            print(f"   After upload to Google Sheets, set data validation on {dropdown_col}{first_row}:{dropdown_col}{last_row} manually.")
        except PermissionError:
            base, ext = os.path.splitext(out_xlsx)
            alt = f"{base}_{_time.strftime('%Y%m%d_%H%M%S')}{ext}"
            wb.save(alt)
            print(f"⚠️ File in use; saved as → {alt}")


# --- Option 2: Export unique suburbs -----------------------------------------
def export_linz_suburbs(
    src_folder: str = "Street database",
    src_file: str = "linz_auckland_addresses.csv",
    out_xlsx: str = "Exported_LINZ_Suburbs.xlsx",
) -> None:
    """
    Load LINZ CSV, extract unique 'Suburb', normalize, sort, export to Excel.
    Shows a progress bar for each stage.
    """
    import os, time
    import pandas as pd
    from contextlib import contextmanager

    @contextmanager
    def _stage_bar(desc: str, total: int = 1, unit: str = "step"):
        bar = _tqdm(total=total, desc=desc, unit=unit, dynamic_ncols=True, leave=False)
        try:
            yield bar
        finally:
            if bar.n < bar.total:
                bar.update(bar.total - bar.n)
            bar.close()

    def _maybe_cancel(_: str) -> bool:
        try:
            return "cancel_flag" in globals() and cancel_flag.is_set()
        except Exception:
            return False

    # 1) Locate source CSV
    with _stage_bar("📁 Locate Source CSV") as bar:
        path = os.path.join(src_folder, src_file)
        if not os.path.exists(path):
            if os.path.exists(src_file):
                path = src_file
            else:
                print(f"❌ Could not find source CSV at '{path}' or '{src_file}'.")
                return
        bar.update(1)
    if _maybe_cancel("Locate Source CSV"):
        print("⚠️  Cancelled during: Locate Source CSV"); return

    # 2) Load CSV
    with _stage_bar("📥 Load CSV") as bar:
        try:
            df = pd.read_csv(path, dtype=str, encoding="utf-8-sig")
        except Exception:
            df = pd.read_csv(path, dtype=str)
        df = df.fillna("")
        if "Suburb" not in df.columns:
            print(f"❌ Missing required column 'Suburb' in {path}")
            return
        bar.update(1)
    if _maybe_cancel("Load CSV"):
        print("⚠️  Cancelled during: Load CSV"); return

    # 3) Clean + unique suburbs
    with _stage_bar("🧹 Extract Unique Suburbs") as bar:
        series = (
            df["Suburb"]
            .astype(str)
            .str.replace(r"\s+", " ", regex=True)
            .str.strip()
            .str.title()
        )
        series = series[series != ""]
        unique_suburbs = sorted(series.unique().tolist())
        if not unique_suburbs:
            print("ℹ️ No suburbs found — nothing to export."); return
        bar.update(1)
    if _maybe_cancel("Extract Unique Suburbs"):
        print("⚠️  Cancelled during: Extract Unique Suburbs"); return

    # 4) Write Excel
    with _stage_bar("💾 Write Excel") as bar:
        out_dir = src_folder if os.path.isdir(src_folder) else "."
        out_path = os.path.join(out_dir, out_xlsx)
        try:
            pd.DataFrame({"Suburb": unique_suburbs}).to_excel(
                out_path, sheet_name="Suburbs", index=False
            )
            print(f"✅ Exported {len(unique_suburbs)} unique suburbs → {out_path}")
            bar.update(1)
        except PermissionError:
            base, ext = os.path.splitext(out_xlsx)
            ts = time.strftime("%Y%m%d_%H%M%S")
            alt_path = os.path.join(out_dir, f"{base}_{ts}{ext}")
            pd.DataFrame({"Suburb": unique_suburbs}).to_excel(
                alt_path, sheet_name="Suburbs", index=False
            )
            print(f"⚠️  File in use; saved as → {alt_path}")
            bar.update(1)
        except Exception as e:
            print(f"❌ Failed to write Excel: {e}")


# --- Option 1: Export LINZ Streets -------------------------------------------
def export_linz_streets(
    src_folder: str = "Street database",
    src_file: str = "linz_auckland_addresses.csv",
    out_xlsx: str = "Exported_LINZ_Streets.xlsx",
) -> None:
    """
    Load a LINZ 'all addresses' CSV, filter by allow-list, extract 'Street, Suburb'
    (stripping unit/house prefixes), de-dupe + sort, write to Excel.
    Shows a progress bar for each stage.
    """
    import os, re, time
    import pandas as pd
    from contextlib import contextmanager

    @contextmanager
    def _stage_bar(desc: str, total: int = 1, unit: str = "step"):
        bar = _tqdm(total=total, desc=desc, unit=unit, dynamic_ncols=True, leave=False)
        try:
            yield bar
        finally:
            if bar.n < bar.total:
                bar.update(bar.total - bar.n)
            bar.close()

    def _maybe_cancel(_: str) -> bool:
        try:
            return "cancel_flag" in globals() and cancel_flag.is_set()
        except Exception:
            return False

    allowed_suburbs_raw = [
        "Airport Oaks","Alfriston","Ardmore","Auckland Airport","Beachlands","Bombay","Bombay Hills","Botany Downs",
        "Bucklands Beach","Burswood","Clendon Park","Clevedon","Clover Park","Cockle Bay","Conifer Grove","Dannemora",
        "Drury","East Tāmaki","East Tāmaki Heights","Eastern Beach","Ellerslie","Epsom","Farm Cove","Favona",
        "Flat Bush","Glen Eden","Glen Innes","Glenbrook","Glendene","Glendowie","Golflands","Goodwood Heights",
        "Greenlane","Half Moon Bay","Highland Park","Hillpark","Howick","Hunua","Huntington Park","Karaka",
        "Kawakawa Bay","Kohimarama","Mangere","Mangere Bridge","Manukau","Manukau Harbour","Manukau Heads","Manurewa",
        "Manurewa East","Maraetai","Meadowbank","Mellons Bay","Mission Bay","Mount Eden","Mount Wellington","Māngere",
        "Māngere Bridge","Māngere East","Northpark","Onehunga","Orakei","Ormiston","Otahuhu","Pahurehure",
        "Pakuranga","Pakuranga Heights","Panmure","Papakura","Papatoetoe","Penrose","Point Chevalier","Point England",
        "Pukekohe","Remuera","Saint Heliers","Saint Johns","Shamrock Park","Shelly Park","Somerville","St Johns",
        "Sunnyhills","Takanini","Totara Heights","Wai O Taiki Bay","Waiheke","Wattle Downs","Whitford","Wiri",
    ]
    allowed = {s.casefold() for s in allowed_suburbs_raw}

    # 1) Locate source CSV
    with _stage_bar("📁 Locate Source CSV") as bar:
        path = os.path.join(src_folder, src_file)
        if not os.path.exists(path):
            if os.path.exists(src_file):
                path = src_file
            else:
                print(f"❌ Could not find source CSV at '{path}' or '{src_file}'.")
                return
        bar.update(1)
    if _maybe_cancel("Locate Source CSV"):
        print("⚠️  Cancelled during: Locate Source CSV"); return

    # 2) Load CSV
    with _stage_bar("📥 Load CSV") as bar:
        try:
            df = pd.read_csv(path, dtype=str, encoding="utf-8-sig")
        except Exception:
            df = pd.read_csv(path, dtype=str)
        df = df.fillna("")
        for col in ("full_add_2", "Suburb"):
            if col not in df.columns:
                print(f"❌ Missing required column '{col}' in {path}")
                return
        bar.update(1)
    if _maybe_cancel("Load CSV"):
        print("⚠️  Cancelled during: Load CSV"); return

    # 3) Filter by Suburb allowlist (fast vectorized op → single-step bar)
    with _stage_bar("🧮 Filter by Allowed Suburbs") as bar:
        mask = df["Suburb"].astype(str).map(lambda s: s.casefold() in allowed)
        df_allowed = df.loc[mask].copy()
        bar.update(1)
    if _maybe_cancel("Filter by Allowed Suburbs"):
        print("⚠️  Cancelled during: Filter by Allowed Suburbs"); return

    # 4) Extract "Street, Suburb" (per-row)
    _LEADING_PREFIX_RE = re.compile(
        r"""
        ^\s*
        (?:
            (?:Unit\s*[A-Za-z0-9]+/)?           # optional "Unit 3/"
            (?=[A-Za-z0-9/-]*\d)
            [A-Za-z0-9]+(?:[/-][A-Za-z0-9]+)*
        )
        \s+
        """,
        re.IGNORECASE | re.VERBOSE,
    )

    def _extract_street_suburb(full_add_2: str) -> str | None:
        if not full_add_2:
            return None
        parts = [p.strip() for p in str(full_add_2).split(",")]
        if len(parts) < 2:
            return None
        street_part, suburb_part = parts[0], parts[1]
        street_part = _LEADING_PREFIX_RE.sub("", street_part).strip()
        street_part = " ".join(street_part.split()).title()
        suburb_part = " ".join(suburb_part.split()).title()
        if not street_part or not suburb_part:
            return None
        return f"{street_part}, {suburb_part}"

    extracted: list[str] = []
    if len(df_allowed):
        for val in _tqdm(
            df_allowed["full_add_2"].astype(str),
            desc="🔎 Extract Street, Suburb",
            unit="row",
            dynamic_ncols=True,
            leave=False,
        ):
            if _maybe_cancel("Extract Street, Suburb"):
                print("⚠️  Cancelled during: Extract Street, Suburb"); return
            item = _extract_street_suburb(val)
            if item:
                extracted.append(item)

    # 5) De-duplicate + sort
    with _stage_bar("🧹 De-duplicate + Sort") as bar:
        unique_sorted = sorted(set(extracted))
        if not unique_sorted:
            print("ℹ️ No streets matched the criteria — nothing to export."); return
        bar.update(1)
    if _maybe_cancel("De-duplicate + Sort"):
        print("⚠️  Cancelled during: De-duplicate + Sort"); return

    # 6) Write Excel
    with _stage_bar("💾 Write Excel") as bar:
        out_dir = src_folder if os.path.isdir(src_folder) else "."
        out_path = os.path.join(out_dir, out_xlsx)
        try:
            pd.DataFrame({"Street, Suburb": unique_sorted}).to_excel(
                out_path, sheet_name="Streets", index=False
            )
            print(f"✅ Exported {len(unique_sorted)} unique streets → {out_path}")
            bar.update(1)
        except PermissionError:
            base, ext = os.path.splitext(out_xlsx)
            ts = time.strftime("%Y%m%d_%H%M%S")
            alt_path = os.path.join(out_dir, f"{base}_{ts}{ext}")
            pd.DataFrame({"Street, Suburb": unique_sorted}).to_excel(
                alt_path, sheet_name="Streets", index=False
            )
            print(f"⚠️  File in use; saved as → {alt_path}")
            bar.update(1)
        except Exception as e:
            print(f"❌ Failed to write Excel: {e}")



# --- Option 5: Find Missing Addresses ---------------------------------------
def find_missing_addresses(
    src_folder: str = "Street database",
    src_file: str = "Missing_Streets.csv",
    out_input_csv: str = "input_nws.csv",
) -> None:
    """
    Option 5 (reworked):
      • Compare Street_all vs Street_now to get missing (Street, Suburb) pairs.
      • Probe numbers 1 → 2..5 → 6..10; accept first in-Auckland hit and seed coords.
      • Keep ALL (Street, Suburb) variants; set guard flags for multi-suburb streets.
      • Write input_nws.csv.
      • Run Option 4 but monkey-patch heavy pre-clean stages to NO-OPs so we de-dupe → geocode.
    """
    import os, csv, sys
    import pandas as pd
    from contextlib import contextmanager
    from collections import defaultdict

    @contextmanager
    def _stage(msg: str):
        print(f"▶ {msg} ...", end="", flush=True)
        try:
            yield
            print(" done.")
        except Exception:
            print(" failed.")
            raise

    # ---------- helpers ----------
    def _fmt_addr(number: str, street: str, suburb: str) -> str:
        if "fmt_addr_parts" in globals():
            try:
                return globals()["fmt_addr_parts"](number, street, suburb or "Auckland")
            except Exception:
                pass
        n = (number or "").strip()
        s = (street or "").strip()
        sb = (suburb or "").strip() or "Auckland"
        return f"{n} {s}, {sb}, Auckland".strip(", ")

    def _is_in_akl(lat, lon) -> bool:
        if "is_in_auckland" in globals():
            try:
                return bool(globals()["is_in_auckland"](float(lat), float(lon)))
            except Exception:
                return False
        try:
            la, lo = float(lat), float(lon)
            return (-37.30 <= la <= -36.20) and (174.30 <= lo <= 175.60)
        except Exception:
            return False

    def _canon_suburb(s: str) -> str:
        s = (s or "").strip().title()
        if not s:
            return ""
        if "macron_suburb_map" in globals():
            return globals()["macron_suburb_map"].get(s, s)
        return s

    def _hit_ok(t) -> bool:
        if not (isinstance(t, tuple) and len(t) == 4):
            return False
        try:
            la = float(t[1]); lo = float(t[2])
        except Exception:
            return False
        return _is_in_akl(la, lo)

    def _geocode(addr: str):
        fn = globals().get("get_lat_long")
        if not callable(fn):
            return None
        try:
            return fn(addr)
        except Exception:
            return None

    def _probe_number(street: str, suburb: str):
        for n in [1,2,3,4,5,6,7,8,9,10]:
            res = _geocode(_fmt_addr(str(n), street, suburb))
            if _hit_ok(res):
                return str(n), res
        return "1", None

    # ---------- surgical skip (patch the RIGHT module) ----------
    @contextmanager
    def _surgical_skip_stages():
        """
        Patch the module that owns process_csv/run_clean_verify_and_split_after_purge.
        This ensures 3.3/3.4/3.5(pre)/3.6 really don't run.
        """
        import types

        # Locate the owner module (prefer the one that contains process_csv)
        owner_mod = None
        try:
            if "process_csv" in globals() and isinstance(globals()["process_csv"], types.FunctionType):
                owner_mod = sys.modules[globals()["process_csv"].__module__]
            elif "run_clean_verify_and_split_after_purge" in globals() and isinstance(globals()["run_clean_verify_and_split_after_purge"], types.FunctionType):
                owner_mod = sys.modules[globals()["run_clean_verify_and_split_after_purge"].__module__]
            else:
                # fallback import by name if available
                owner_mod = __import__("Clean_NewWorldScheduler")
        except Exception:
            owner_mod = None

        if owner_mod is None:
            # If we can't find the module, do nothing but still let the flow continue.
            yield
            return

        # no-op shims
        def _noop_33(all_rows, *_, **__):
            print("   ↪ SKIP 3.3 pre_correct_street_spellings"); return all_rows, 0
        def _noop_34(all_rows, *_, **__):
            print("   ↪ SKIP 3.4 standardize_similar_streets"); return all_rows, 0
        def _noop_resolve(*_, **__):
            print("   ↪ SKIP 3.5 resolve_conflicting_suburbs_by_proximity (pre)"); return None
        def _noop_finalnorm(all_rows, *_, **__):
            print("   ↪ SKIP 3.6 final_normalize_rows"); return all_rows, 0
        def _identity(rows, *_, **__):
            print("   ↪ SKIP 3.6 enforce_final_street_spelling"); return rows

        targets = {
            "pre_correct_street_spellings": _noop_33,              # 3.3
            "standardize_similar_streets": _noop_34,               # 3.4
            "resolve_conflicting_suburbs_by_proximity": _noop_resolve,  # 3.5 (this disables both pre/post;
                                                                        # if you want ONLY pre skipped,
                                                                        # add a phase flag and check it here)
            "final_normalize_rows": _noop_finalnorm,               # 3.6a
            "enforce_final_street_spelling": _identity,            # 3.6b
        }

        originals = {}
        try:
            for name, repl in targets.items():
                if hasattr(owner_mod, name) and callable(getattr(owner_mod, name)):
                    originals[name] = getattr(owner_mod, name)
                    setattr(owner_mod, name, repl)
            yield
        finally:
            for name, fn in originals.items():
                setattr(owner_mod, name, fn)

    # ---------- 1) Load Missing_Streets.csv ----------
    with _stage("Locate and load Missing_Streets.csv"):
        path = os.path.join(src_folder, src_file)
        if not os.path.exists(path):
            print(f"\n❌ Could not find '{src_file}' in {src_folder}")
            return
        try:
            df = pd.read_csv(path, dtype=str, encoding="utf-8-sig")
        except Exception:
            df = pd.read_csv(path, dtype=str)
        df = df.fillna("")
        if "Street_all" not in df.columns or "Street_now" not in df.columns:
            print("\n❌ Missing required columns: 'Street_all' and 'Street_now'")
            return

    # ---------- 2) Compute missing ----------
    with _stage("Compute missing streets"):
        all_records = []
        for val in df["Street_all"].astype(str):
            if "," not in val:
                continue
            st, sb = [p.strip() for p in val.split(",", 1)]
            if st and sb:
                all_records.append((st, sb))
        all_df = pd.DataFrame(all_records, columns=["Street", "Suburb"]).drop_duplicates()
        now_streets = set(df["Street_now"].astype(str).str.strip())
        missing_df = all_df[~all_df["Street"].isin(now_streets)].drop_duplicates()
        if missing_df.empty:
            print("\nℹ️ No missing streets found — nothing to do.")
            return

        # streets with multiple suburbs → keep all variants and guard later stages
        street_to_suburbs = defaultdict(set)
        for _, r in missing_df.iterrows():
            street_to_suburbs[(r["Street"] or "").strip().title()].add((r["Suburb"] or "").strip().title())
        multi_suburb_streets = {st for st, subs in street_to_suburbs.items() if len(subs) > 1}
        try:
            if "PROTECTED_STREETS" in globals() and isinstance(globals()["PROTECTED_STREETS"], set):
                globals()["PROTECTED_STREETS"] |= set(multi_suburb_streets)
        except Exception:
            pass

    # ---------- 3) Probe + build input_nws.csv rows ----------
    with _stage(f"Probe numbers & seed coords for {len(missing_df)} streets"):
        out_rows = []
        for _, row in missing_df.iterrows():
            street = (row["Street"] or "").strip().title()
            suburb = _canon_suburb((row["Suburb"] or "").strip())

            number, geo = _probe_number(street, suburb)

            lat = lon = postal = ""
            final_suburb = suburb
            if geo:
                label, la, lo, pc = geo
                parts = [p.strip() for p in (label or "").split(",")]
                if len(parts) > 1:
                    g_suburb = _canon_suburb(parts[1])
                    if g_suburb:
                        final_suburb = g_suburb
                lat = f"{float(la):.8f}"
                lon = f"{float(lo):.8f}"
                postal = pc or ""

            flags = {}
            if street in multi_suburb_streets:
                flags["_skip_suburb_lookup"] = "1"
                flags["_lock_suburb"] = "1"

            out_rows.append({
                "Number": number,
                "Street": street,
                "Suburb": final_suburb,
                "PostalCode": postal,
                "Status": "",
                "Latitude": lat,
                "Longitude": lon,
                **flags,
            })

    # ---------- 4) Write input_nws.csv ----------
    with _stage(f"Write {out_input_csv}"):
        fieldnames = ["Number","Street","Suburb","PostalCode","Status","Latitude","Longitude",
                      "_skip_suburb_lookup","_lock_suburb"]
        for r in out_rows:
            for k in fieldnames:
                r.setdefault(k, "")
        with open(out_input_csv, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore", restval="")
            w.writeheader()
            w.writerows(out_rows)
        print(f"\n✅ Prepared {len(out_rows)} row(s) in '{out_input_csv}'")

    # ---------- 5) Run Option 4 with heavy stages disabled ----------
    print("▶ Run Clean_NewWorldScheduler Option 4 (Duplicates → Geocode) ...")
    try:
        with _surgical_skip_stages():
            if "run_clean_verify_and_split_after_purge" in globals():
                globals()["run_clean_verify_and_split_after_purge"]()
            else:
                from Clean_NewWorldScheduler import run_clean_verify_and_split_after_purge as _run_opt4
                _run_opt4()
        print("✅ Option 4 finished.")
    except Exception as e:
        print(f"⚠️ Could not execute Option 4 automatically: {e}")
        print("   You can run it manually afterwards.")





# --- Option 10 submenu -------------------------------------------------------
def finding_new_addresses() -> None:
    """
    Simple numeric submenu:
      1) Export Streets  → Excel
      2) Export Suburbs  → Excel
      3) Clean Exported_LINZ_Streets File → Excel
      4) Create Excel For New Addresses (E18..E530 with fixed Z-block ranges)
      5) Find Missing Addresses → CSVs split per suburb
      0) Back
    """
    # Defaults
    src_folder = "Street database"
    src_file   = "linz_auckland_addresses.csv"
    out_streets_xlsx       = "Exported_LINZ_Streets.xlsx"
    out_streets_clean_xlsx = "Exported_LINZ_Streets_Clean.xlsx"
    out_suburbs_xlsx       = "Exported_LINZ_Suburbs.xlsx"
    out_new_addresses_xlsx = "New_Addresses_Template.xlsx"   # for Option 4

    def _print_header() -> None:
        print("\n" + "=" * 68)
        print("  Option 10 — Finding New Addresses")
        print("=" * 68)
        print(f"  Source folder : {src_folder}")
        print(f"  Source file   : {src_file}")
        print("-" * 68)
        print("  1- Export Streets (linz_auckland_addresses.csv - Exported_LINZ_Streets.csv)")
        print("  2- Export Suburbs (linz_auckland_addresses.csv - Exported_LINZ_Suburbs.csv.csv)")
        print("  3- Clean Exported_LINZ_Streets File (clean_exported_linz_streets - Exported_LINZ_Streets_Clean.csv)")
        print("  4- Create Excel For New Addresses (create_excel_for_new_addresses - Exported_LINZ_Streets_Clean.xlsx)")
        print("  5- Find Missing Addresses (find_missing_addresses - missing_addresses.csv)")
        print("  0- Back")
        print("-" * 68)

    def _clear_cancel():
        if "cancel_flag" in globals() and hasattr(globals()["cancel_flag"], "clear"):
            try:
                globals()["cancel_flag"].clear()
            except Exception:
                pass

    def _do_export_streets() -> None:
        _clear_cancel()
        export_linz_streets(src_folder=src_folder, src_file=src_file, out_xlsx=out_streets_xlsx)

    def _do_export_suburbs() -> None:
        _clear_cancel()
        export_linz_suburbs(src_folder=src_folder, src_file=src_file, out_xlsx=out_suburbs_xlsx)

    def _do_clean_streets() -> None:
        _clear_cancel()
        clean_exported_linz_streets(src_folder=src_folder, in_xlsx=out_streets_xlsx, out_xlsx=out_streets_clean_xlsx)

    def _do_create_excel_new_addresses() -> None:
        _clear_cancel()
        # Simple segmented method: E18..E530 each uses its own 500-row block in Z.
        create_excel_for_new_addresses(
            out_xlsx=out_new_addresses_xlsx,
            sheet_name="New Addresses",
            first_row=18,
            num_rows=530 - 18 + 1,   # E18..E530 (513 rows)
            block_size=500,
            # hide_helpers=True,     # optional: hide column Z
        )

    def _do_find_missing_addresses() -> None:
        _clear_cancel()
        find_missing_addresses(src_folder=src_folder, src_file="Missing_Streets.csv")

    while True:
        _print_header()
        choice = (input("Choose (0-5): ") or "").strip().lower()

        if choice in {"0", "q", "z"}:
            print("↩ Back to main menu.")
            return

        elif choice == "1":
            try:
                if "run_with_cancel" in globals():
                    globals()["run_with_cancel"](_do_export_streets)
                else:
                    _do_export_streets()
            except KeyboardInterrupt:
                print("\n⚠️  Cancelled by user.")
            except PermissionError as e:
                print(f"❌ Failed to write Excel (file locked?): {e}")
            except Exception as e:
                print(f"❌ Export error: {e}")

        elif choice == "2":
            try:
                if "run_with_cancel" in globals():
                    globals()["run_with_cancel"](_do_export_suburbs)
                else:
                    _do_export_suburbs()
            except KeyboardInterrupt:
                print("\n⚠️  Cancelled by user.")
            except PermissionError as e:
                print(f"❌ Failed to write Excel (file locked?): {e}")
            except Exception as e:
                print(f"❌ Export error: {e}")

        elif choice == "3":
            try:
                if "run_with_cancel" in globals():
                    globals()["run_with_cancel"](_do_clean_streets)
                else:
                    _do_clean_streets()
            except KeyboardInterrupt:
                print("\n⚠️  Cancelled by user.")
            except PermissionError as e:
                print(f"❌ Failed to write Excel (file locked?): {e}")
            except Exception as e:
                print(f"❌ Clean error: {e}")

        elif choice == "4":
            try:
                if "run_with_cancel" in globals():
                    globals()["run_with_cancel"](_do_create_excel_new_addresses)
                else:
                    _do_create_excel_new_addresses()
            except KeyboardInterrupt:
                print("\n⚠️  Cancelled by user.")
            except PermissionError as e:
                print(f"❌ Failed to write Excel (file locked?): {e}")
            except Exception as e:
                print(f"❌ Create Excel error: {e}")

        elif choice == "5":
            try:
                if "run_with_cancel" in globals():
                    globals()["run_with_cancel"](_do_find_missing_addresses)
                else:
                    _do_find_missing_addresses()
            except KeyboardInterrupt:
                print("\n⚠️  Cancelled by user.")
            except Exception as e:
                print(f"❌ Missing addresses error: {e}")

        else:
            print("⚠️  Invalid choice. Please pick 0–5.")



# Optional: allow running this module standalone
if __name__ == "__main__":
    finding_new_addresses()
